import openpyxl
from django.shortcuts import redirect, render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from .models import *


def home(request):
    return render(request, 'home.html')

from django.db import transaction
import openpyxl

def import_student_scores(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']

        with transaction.atomic():
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            subjects_data = [
                ("اللغة العربية", 50),
                ("اللغة الإنجليزية", 50),
                ("MATH", 50),
                ("PHYSICS", 40),
                ("تخصص نظري", 130),
                ("تخصص عملي", 130),
                ("التدريب الميداني", 100),
                ("المجموع", 550),
                ("التربية الدينية", 20),
                ("التربية الوطنية", 20)
            ]

            existing_subjects = {subject.name: subject for subject in Subject.objects.all()}

            subjects_to_create = [
                Subject(name=subject_name, max_score=max_score)
                for subject_name, max_score in subjects_data
                if subject_name not in existing_subjects
            ]

            if subjects_to_create:
                Subject.objects.bulk_create(subjects_to_create)

            existing_subjects.update({
                subject.name: subject
                for subject in Subject.objects.filter(name__in=[subj[0] for subj in subjects_data])
            })

            scores_data = []

            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
                try:
                    national_id = row[0]
                    full_name = row[1]

                    subjects_scores = [
                        ("اللغة العربية", row[2]),
                        ("اللغة الإنجليزية", row[3]),
                        ("MATH", row[4]),
                        ("PHYSICS", row[5]),
                        ("تخصص نظري", row[6]),
                        ("تخصص عملي", row[7]),
                        ("التدريب الميداني", row[8]),
                        ("المجموع", row[9]),
                        ("التربية الدينية", row[10]),
                        ("التربية الوطنية", row[11])
                    ]

                    if not national_id or not full_name:
                        continue  # تخطي الصفوف غير الكاملة

                    student, created = Student.objects.get_or_create(
                        national_no=national_id,
                        defaults={'name': full_name}
                    )

                    for subject_name, score in subjects_scores:
                        subject = existing_subjects.get(subject_name)
                        try:
                            score = float(score)
                        except (TypeError, ValueError):
                            continue  # تجاهل القيمة إذا لم تكن رقمية

                        if subject:
                            scores_data.append(Score(student=student, subject=subject, score=score))

                except Exception as e:
                    print("خطأ في الصف:", row)
                    print("الخطأ:", e)
                    continue

            if scores_data:
                Score.objects.bulk_create(scores_data)

        return render(request, 'student_excel.html', {'message': 'تم رفع الملف بنجاح'})

    return render(request, 'student_excel.html')

def student_score_view(request):
    score_data = None
    student_name = None  
    student_id = None
    student_percentage = None  
    fail_count = 0

    if request.method == 'GET':
        national_id = request.GET.get('national_id', None)
        
        if national_id:
            try:
                student = Student.objects.get(national_no=national_id)
                student_name = student.name  
                student_id = student.national_no

                scores = Score.objects.filter(student=student).select_related('subject')

                total_score = 0
                max_total_score = 0
                score_data = {}

                for score in scores:
                    subject_max = score.subject.max_score
                    percentage = (score.score / subject_max) * 100 if subject_max > 0 else 0
                    failed = percentage < 50

                    if failed:
                        fail_count += 1
                    
                    score_data[score.subject.name] = {
                        'score': score.score,
                        'max_score': subject_max,
                        'percentage': percentage,
                        'failed': failed
                    }

                    total_score += score.score
                    max_total_score += subject_max

                if max_total_score > 0:
                    student_percentage = (total_score / max_total_score) * 100

            except Student.DoesNotExist:
                score_data = "الطالب غير موجود برقم القومي المدخل."
        
    return render(request, 'student_score.html', {
        'score_data': score_data,
        'student_name': student_name,
        'student_id': student_id,
        'student_percentage': student_percentage,
        'fail_count': fail_count
    })

